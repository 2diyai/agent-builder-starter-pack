from docling.document_converter import DocumentConverter
from docling.datamodel.base_models import DocumentStream
from docling_core.transforms.serializer.markdown import MarkdownDocSerializer
from docling_core.transforms.serializer.base import SerializationResult
from docling_core.transforms.serializer.common import create_ser_result
from docling_core.types.doc.document import DoclingDocument, GroupItem
from fastapi import FastAPI, File, HTTPException, Query, UploadFile, Request
from io import BytesIO
from openpyxl import load_workbook
from pydantic import BaseModel
from typing import Any, Optional

from app.excel_pipeline import ExcelToLLMPipeline, JsonFormat

app = FastAPI()

class Payload(BaseModel):
    text: str

@app.get("/health")
def health():
    return {"status": "ok"}

@app.post("/test-post")
def run(payload: Payload):
    # Minimal example: reverse the input text
    return {"result": f"{payload.text} -> {payload.text[::-1]}"}

class MultiplyPayload(BaseModel):
    a: float
    b: float


SUPPORTED_OFFICE_EXTENSIONS: dict[str, str] = {
    ".xlsx": "excel",
    ".xlsm": "excel",
    ".xltx": "excel",
    ".xltm": "excel",
    ".docx": "word",
    ".docm": "word",
    ".dotx": "word",
    ".dotm": "word",
    ".pptx": "powerpoint",
    ".pptm": "powerpoint",
    ".potx": "powerpoint",
    ".potm": "powerpoint",
}


def _ext(name: str) -> str:
    name = (name or "").lower().strip()
    if "." not in name:
        return ""
    return "." + name.rsplit(".", 1)[-1]


def _validate_supported_office_filename(filename: str) -> None:
    if _ext(filename) not in SUPPORTED_OFFICE_EXTENSIONS:
        allowed = ", ".join(sorted(SUPPORTED_OFFICE_EXTENSIONS.keys()))
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type. Allowed extensions: {allowed}",
        )


def _infer_doc_type(filename: str) -> str:
    return SUPPORTED_OFFICE_EXTENSIONS.get(_ext(filename), "document")


async def _extract_file_bytes(
    request: Request,
    fallback_filename: str,
) -> tuple[bytes, str]:
    content_type = (request.headers.get("content-type") or "").lower()

    if "multipart/form-data" in content_type:
        form = await request.form()
        for _, value in form.multi_items():
            if isinstance(value, UploadFile):
                filename = value.filename or fallback_filename
                file_bytes = await value.read()
                return file_bytes, filename
        raise HTTPException(status_code=400, detail="No file found in multipart form-data")

    file_bytes = await request.body()
    return file_bytes, fallback_filename


class HierarchyMarkdownDocSerializer(MarkdownDocSerializer):
    def _format_group(self, grp: GroupItem) -> str:
        label = getattr(grp, "label", None)
        label_txt = label.value if hasattr(label, "value") else str(label or "GROUP")
        name = getattr(grp, "name", None) or getattr(grp, "text", None)
        return f"{label_txt}:{name}" if name else label_txt

    def _meta_start(self, group_id: str) -> SerializationResult:
        return create_ser_result(text=f'<meta id="{group_id}">')

    def _meta_end(self) -> SerializationResult:
        return create_ser_result(text="</meta>")

    @staticmethod
    def _common_prefix_len(a: tuple[str, ...], b: tuple[str, ...]) -> int:
        size = min(len(a), len(b))
        idx = 0
        while idx < size and a[idx] == b[idx]:
            idx += 1
        return idx

    def get_parts(
        self,
        item=None,
        *,
        traverse_pictures: bool = False,
        list_level: int = 0,
        is_inline_scope: bool = False,
        visited: Optional[set[str]] = None,
        **kwargs: Any,
    ) -> list[SerializationResult]:
        parts: list[SerializationResult] = []
        my_visited: set[str] = visited if visited is not None else set()
        params = self.params.merge_with_patch(patch=kwargs)
        add_content = True
        if hasattr(params, "add_content"):
            add_content = getattr(params, "add_content")

        active_groups_by_level: dict[int, GroupItem] = {}
        last_path: tuple[str, ...] = ()

        for node, level in self.doc.iterate_items(
            root=item,
            with_groups=True,
            traverse_pictures=params.traverse_pictures,
            included_content_layers=params.layers,
        ):
            if isinstance(node, GroupItem):
                active_groups_by_level[level] = node
                for deeper in [k for k in active_groups_by_level if k > level]:
                    active_groups_by_level.pop(deeper)
                continue

            if node.self_ref in my_visited:
                continue
            my_visited.add(node.self_ref)

            part = super().serialize(
                item=node,
                list_level=list_level,
                is_inline_scope=is_inline_scope,
                visited=my_visited,
                **(dict(level=level) | kwargs),
            )
            if not part.text and add_content:
                continue

            ordered_levels = sorted(k for k in active_groups_by_level if 0 < k < level)
            current_path = tuple(self._format_group(active_groups_by_level[k]) for k in ordered_levels)

            if current_path != last_path:
                cpl = self._common_prefix_len(last_path, current_path)

                for _ in range(len(last_path), cpl, -1):
                    parts.append(self._meta_end())

                for i in range(cpl, len(current_path)):
                    parts.append(self._meta_start(current_path[i]))

                last_path = current_path

            if part.text or not add_content:
                parts.append(part)

        for _ in range(len(last_path), 0, -1):
            parts.append(self._meta_end())

        return parts


def export_to_markdown_with_hierarchy(
    doc: DoclingDocument,
    doc_type: str,
    source: str,
    **kwargs: Any,
) -> str:
    serializer = HierarchyMarkdownDocSerializer(doc=doc)
    body = serializer.serialize(**kwargs).text
    return f'<<DOC_START type={doc_type} source="{source}">>\n{body}\n<<DOC_END>>'

@app.post("/multiply")
def multiply(payload: MultiplyPayload):
    return {"result": payload.a * payload.b}

@app.post("/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    if not file.filename:
        raise HTTPException(status_code=400, detail="Missing file name")

    allowed_extensions = (".xlsx", ".xlsm", ".xltx", ".xltm")
    if not file.filename.lower().endswith(allowed_extensions):
        raise HTTPException(
            status_code=400,
            detail="Only modern Excel files are supported (.xlsx, .xlsm, .xltx, .xltm)",
        )

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    try:
        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    except Exception as error:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {error}")

    sheets = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        sheets.append(
            {
                "name": sheet_name,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
            }
        )

    return {
        "filename": file.filename,
        "sheet_count": len(workbook.sheetnames),
        "sheets": sheets,
    }


@app.post("/upload-excel-structured")
async def upload_excel_structured(
    file: UploadFile = File(...),
    json_format: str = Query(default=JsonFormat.HYBRID, pattern="^(flat|object|hybrid)$"),
    max_rows_per_table: int | None = Query(default=None, ge=1),
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="Missing file name")

    allowed_extensions = (".xlsx", ".xlsm", ".xltx", ".xltm")
    if not file.filename.lower().endswith(allowed_extensions):
        raise HTTPException(
            status_code=400,
            detail="Only modern Excel files are supported (.xlsx, .xlsm, .xltx, .xltm)",
        )

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    try:
        pipeline = ExcelToLLMPipeline(
            json_format=json_format,
            max_rows_per_table=max_rows_per_table,
        )
        return pipeline.process_workbook_bytes(file_bytes, source_name=file.filename)
    except Exception as error:
        raise HTTPException(status_code=400, detail=f"Unable to process workbook: {error}")

@app.post("/docling")
async def docling(
    request: Request,
    filename: str = Query(default="upload.xlsx"),
):
    file_bytes, stream_name = await _extract_file_bytes(request, fallback_filename=filename)

    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    _validate_supported_office_filename(stream_name)

    try:
        doc_stream = DocumentStream(name=stream_name, stream=BytesIO(file_bytes))

        converter = DocumentConverter()
        result = converter.convert(doc_stream)

        markdown = result.document.export_to_markdown()
        return {"markdown": f"The content of the file is:\n{markdown}", "filename": stream_name}
    except Exception as error:
        raise HTTPException(status_code=400, detail=f"Unable to process document: {error}")


@app.post("/docling-hierarchy")
async def docling_hierarchy(
    request: Request,
    filename: str = Query(default="upload.xlsx"),
    doc_type: str | None = Query(default=None),
):
    file_bytes, stream_name = await _extract_file_bytes(request, fallback_filename=filename)

    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    _validate_supported_office_filename(stream_name)
    effective_doc_type = doc_type or _infer_doc_type(stream_name)

    try:
        doc_stream = DocumentStream(name=stream_name, stream=BytesIO(file_bytes))

        converter = DocumentConverter()
        result = converter.convert(doc_stream)

        markdown = export_to_markdown_with_hierarchy(
            result.document,
            doc_type=effective_doc_type,
            source=stream_name,
        )

        return {"markdown": markdown, "filename": stream_name, "doc_type": effective_doc_type}
    except Exception as error:
        raise HTTPException(status_code=400, detail=f"Unable to process document: {error}")