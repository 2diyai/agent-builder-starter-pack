from abc import ABC, abstractmethod
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Dict, List, Optional

from openpyxl.worksheet.worksheet import Worksheet

from .models import SheetAnalysis, SheetType, TableBoundary


class SheetParser(ABC):
    @abstractmethod
    def parse(
        self,
        ws_values: Worksheet,
        analysis: SheetAnalysis,
        ws_formulas: Optional[Worksheet] = None,
    ) -> Dict[str, Any]:
        pass


class EmptyParser(SheetParser):
    def parse(
        self,
        ws_values: Worksheet,
        analysis: SheetAnalysis,
        ws_formulas: Optional[Worksheet] = None,
    ) -> Dict[str, Any]:
        return {"type": "empty", "data": []}


class TabularParser(SheetParser):
    def parse(
        self,
        ws_values: Worksheet,
        analysis: SheetAnalysis,
        ws_formulas: Optional[Worksheet] = None,
    ) -> Dict[str, Any]:
        if not analysis.table_regions:
            return {"type": "tabular", "headers": [], "data": []}

        table = analysis.table_regions[0]
        headers = self._extract_headers(ws_values, table)
        data_start = table.header_row + 1 if table.header_row is not None else table.start_row

        rows: List[Dict[str, Any]] = []
        for row_idx in range(data_start, table.end_row + 1):
            row_data: Dict[str, Any] = {}
            has_values = False

            for header_idx, col_idx in enumerate(range(table.start_col, table.end_col + 1)):
                value_cell = ws_values.cell(row_idx + 1, col_idx + 1)
                formula_cell = ws_formulas.cell(row_idx + 1, col_idx + 1) if ws_formulas else None
                extracted = self._extract_cell_value(value_cell, formula_cell)
                row_data[headers[header_idx]] = extracted
                if extracted["value"] is not None:
                    has_values = True

            if has_values:
                rows.append(row_data)

        return {
            "type": "tabular",
            "structure": {
                "rows": len(rows),
                "columns": len(headers),
                "has_row_labels": table.has_row_labels,
            },
            "headers": headers,
            "data": rows,
        }

    def _extract_headers(self, ws_values: Worksheet, table: TableBoundary) -> List[str]:
        if table.header_row is None:
            return [
                f"Column_{col_idx + 1}"
                for col_idx in range(table.start_col, table.end_col + 1)
            ]

        headers: List[str] = []
        for col_idx in range(table.start_col, table.end_col + 1):
            value = ws_values.cell(table.header_row + 1, col_idx + 1).value
            headers.append(str(value).strip() if value is not None else f"Column_{col_idx + 1}")
        return headers

    def _extract_cell_value(self, value_cell, formula_cell=None) -> Dict[str, Any]:
        value = value_cell.value
        result: Dict[str, Any] = {"value": None, "type": "empty"}

        if value is None:
            if formula_cell and isinstance(formula_cell.value, str) and formula_cell.value.startswith("="):
                result["type"] = "formula"
                result["formula"] = formula_cell.value
            return result

        if value_cell.is_date and isinstance(value, (datetime, date)):
            result["value"] = value.isoformat()
            result["type"] = "datetime"
            return result

        if isinstance(value, bool):
            result["value"] = value
            result["type"] = "boolean"
            return result

        if isinstance(value, int):
            result["value"] = value
            result["type"] = "integer"
            return result

        if isinstance(value, float):
            result["value"] = value
            result["type"] = "number"
            result["precision"] = self._float_precision(value)
            return result

        if isinstance(value, Decimal):
            result["value"] = str(value)
            result["type"] = "decimal"
            return result

        if formula_cell and isinstance(formula_cell.value, str) and formula_cell.value.startswith("="):
            result["value"] = value
            result["type"] = "formula_result"
            result["formula"] = formula_cell.value
            return result

        result["value"] = str(value)
        result["type"] = "string"
        return result

    def _float_precision(self, value: float) -> int:
        as_text = repr(value)
        if "." not in as_text:
            return 0
        return len(as_text.split(".", maxsplit=1)[1].rstrip("0"))


class SemiStructuredParser(SheetParser):
    def parse(
        self,
        ws_values: Worksheet,
        analysis: SheetAnalysis,
        ws_formulas: Optional[Worksheet] = None,
    ) -> Dict[str, Any]:
        result: Dict[str, Any] = {
            "type": "semi_structured",
            "metadata": {},
            "tables": [],
            "annotations": [],
        }

        for region in analysis.metadata_regions:
            start_row, start_col, end_row, end_col = region
            content = self._extract_region_content(ws_values, start_row, start_col, end_row, end_col)

            if not analysis.table_regions:
                result["annotations"].append({"type": "metadata", "content": content})
                continue

            first_table_row = min(table.start_row for table in analysis.table_regions)
            if start_row < first_table_row:
                result["metadata"]["header"] = content
            else:
                result["annotations"].append({"type": "footer", "content": content})

        tabular_parser = TabularParser()
        for table in analysis.table_regions:
            table_analysis = SheetAnalysis(
                sheet_type=SheetType.TABULAR,
                table_regions=[table],
                metadata_regions=[],
                merged_cells=analysis.merged_cells,
                data_density=analysis.data_density,
                max_row=analysis.max_row,
                max_col=analysis.max_col,
            )
            parsed_table = tabular_parser.parse(ws_values, table_analysis, ws_formulas)
            parsed_table["location"] = {
                "start_row": table.start_row + 1,
                "start_col": table.start_col + 1,
                "end_row": table.end_row + 1,
                "end_col": table.end_col + 1,
            }
            result["tables"].append(parsed_table)

        result["merged_cells"] = analysis.merged_cells
        return result

    def _extract_region_content(
        self,
        ws_values: Worksheet,
        start_row: int,
        start_col: int,
        end_row: int,
        end_col: int,
    ) -> List[Dict[str, Any]]:
        content: List[Dict[str, Any]] = []

        for row_idx in range(start_row, end_row + 1):
            cells: List[Dict[str, Any]] = []
            for col_idx in range(start_col, end_col + 1):
                cell = ws_values.cell(row_idx + 1, col_idx + 1)
                if cell.value is not None:
                    cells.append(
                        {
                            "row": row_idx + 1,
                            "column": col_idx + 1,
                            "value": str(cell.value),
                        }
                    )
            if cells:
                content.append({"row": row_idx + 1, "cells": cells})

        return content


class UnstructuredParser(SheetParser):
    def parse(
        self,
        ws_values: Worksheet,
        analysis: SheetAnalysis,
        ws_formulas: Optional[Worksheet] = None,
    ) -> Dict[str, Any]:
        blocks: List[Dict[str, Any]] = []

        for row_idx in range(1, analysis.max_row + 1):
            content: List[Dict[str, Any]] = []
            for col_idx in range(1, analysis.max_col + 1):
                cell = ws_values.cell(row_idx, col_idx)
                if cell.value is not None:
                    content.append(
                        {
                            "row": row_idx,
                            "column": col_idx,
                            "value": str(cell.value),
                            "type": type(cell.value).__name__,
                        }
                    )
            if content:
                blocks.append({"row": row_idx, "content": content})

        return {
            "type": "unstructured",
            "layout": "document",
            "merged_cells": analysis.merged_cells,
            "blocks": blocks,
        }


class ParserFactory:
    @staticmethod
    def get_parser(sheet_type: SheetType) -> SheetParser:
        parsers = {
            SheetType.TABULAR: TabularParser(),
            SheetType.SEMI_STRUCTURED: SemiStructuredParser(),
            SheetType.UNSTRUCTURED: UnstructuredParser(),
            SheetType.EMPTY: EmptyParser(),
        }
        return parsers[sheet_type]
