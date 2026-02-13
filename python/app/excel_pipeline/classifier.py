from typing import List, Optional, Tuple

from openpyxl.worksheet.worksheet import Worksheet

from .models import SheetAnalysis, SheetType, TableBoundary


class SheetClassifier:
    def __init__(self, density_threshold: float = 0.15):
        self.density_threshold = density_threshold

    def analyze(self, ws: Worksheet) -> SheetAnalysis:
        max_row, max_col = self._get_actual_dimensions(ws)

        if max_row == 0 or max_col == 0:
            return SheetAnalysis(
                sheet_type=SheetType.EMPTY,
                table_regions=[],
                metadata_regions=[],
                merged_cells=[],
                data_density=0.0,
                max_row=0,
                max_col=0,
            )

        cell_matrix = self._build_cell_matrix(ws, max_row, max_col)
        density = self._calculate_density(cell_matrix)
        merged_cells = [str(rng) for rng in ws.merged_cells.ranges]
        table_regions = self._detect_tables(ws, cell_matrix, max_row, max_col)
        sheet_type = self._classify_structure(cell_matrix, table_regions, merged_cells)
        metadata_regions = self._identify_metadata_regions(
            cell_matrix, table_regions, max_row, max_col
        )

        return SheetAnalysis(
            sheet_type=sheet_type,
            table_regions=table_regions,
            metadata_regions=metadata_regions,
            merged_cells=merged_cells,
            data_density=density,
            max_row=max_row,
            max_col=max_col,
        )

    def _get_actual_dimensions(self, ws: Worksheet) -> Tuple[int, int]:
        max_row = 0
        max_col = 0

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    max_row = max(max_row, cell.row)
                    max_col = max(max_col, cell.column)

        return max_row, max_col

    def _build_cell_matrix(self, ws: Worksheet, max_row: int, max_col: int) -> List[List[int]]:
        matrix = [[0 for _ in range(max_col)] for _ in range(max_row)]

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(r, c)
                if cell.value is not None:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        matrix[r - 1][c - 1] = 2
                    else:
                        matrix[r - 1][c - 1] = 1

        for merged_range in ws.merged_cells.ranges:
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    if row <= max_row and col <= max_col:
                        matrix[row - 1][col - 1] = 3

        return matrix

    def _calculate_density(self, cell_matrix: List[List[int]]) -> float:
        total_cells = len(cell_matrix) * len(cell_matrix[0]) if cell_matrix else 0
        if total_cells == 0:
            return 0.0

        non_empty = 0
        for row in cell_matrix:
            for value in row:
                if value != 0:
                    non_empty += 1

        return non_empty / total_cells

    def _detect_tables(
        self,
        ws: Worksheet,
        cell_matrix: List[List[int]],
        max_row: int,
        max_col: int,
    ) -> List[TableBoundary]:
        tables: List[TableBoundary] = []
        row_densities: List[float] = []

        for row in cell_matrix:
            non_empty = sum(1 for value in row if value != 0)
            row_densities.append(non_empty / max_col if max_col else 0)

        in_table = False
        table_start: Optional[int] = None

        for idx, density in enumerate(row_densities):
            if density >= self.density_threshold and not in_table:
                table_start = idx
                in_table = True
            elif density < self.density_threshold and in_table and table_start is not None:
                table = self._refine_table_boundary(
                    ws,
                    cell_matrix,
                    table_start,
                    idx - 1,
                    max_col,
                )
                if table:
                    tables.append(table)
                in_table = False
                table_start = None

        if in_table and table_start is not None:
            table = self._refine_table_boundary(
                ws,
                cell_matrix,
                table_start,
                max_row - 1,
                max_col,
            )
            if table:
                tables.append(table)

        return tables

    def _refine_table_boundary(
        self,
        ws: Worksheet,
        cell_matrix: List[List[int]],
        start_row: int,
        end_row: int,
        max_col: int,
    ) -> Optional[TableBoundary]:
        if start_row > end_row:
            return None

        occupied_cols: List[int] = []
        for col in range(max_col):
            has_data = False
            for row in range(start_row, end_row + 1):
                if cell_matrix[row][col] != 0:
                    has_data = True
                    break
            if has_data:
                occupied_cols.append(col)

        if not occupied_cols:
            return None

        start_col = occupied_cols[0]
        end_col = occupied_cols[-1]

        header_row = None
        for row_idx in range(start_row, min(start_row + 3, end_row + 1)):
            row_values = [
                ws.cell(row_idx + 1, col_idx + 1).value
                for col_idx in range(start_col, end_col + 1)
            ]
            non_empty = [value for value in row_values if value is not None]
            if non_empty and all(isinstance(value, str) for value in non_empty):
                header_row = row_idx
                break

        has_row_labels = False
        if header_row is not None and end_row > header_row + 1:
            first_col_values = [
                ws.cell(row_idx + 1, start_col + 1).value
                for row_idx in range(header_row + 1, end_row + 1)
            ]
            non_empty = [value for value in first_col_values if value is not None]
            if non_empty:
                text_values = sum(1 for value in non_empty if isinstance(value, str))
                has_row_labels = (text_values / len(non_empty)) > 0.7

        return TableBoundary(
            start_row=start_row,
            start_col=start_col,
            end_row=end_row,
            end_col=end_col,
            header_row=header_row,
            has_row_labels=has_row_labels,
            confidence=1.0,
        )

    def _classify_structure(
        self,
        cell_matrix: List[List[int]],
        table_regions: List[TableBoundary],
        merged_cells: List[str],
    ) -> SheetType:
        if not table_regions:
            return SheetType.UNSTRUCTURED

        total_cells = len(cell_matrix) * len(cell_matrix[0]) if cell_matrix else 0
        main_table = max(
            table_regions,
            key=lambda table: (table.end_row - table.start_row + 1)
            * (table.end_col - table.start_col + 1),
        )
        table_cells = (main_table.end_row - main_table.start_row + 1) * (
            main_table.end_col - main_table.start_col + 1
        )
        coverage = table_cells / total_cells if total_cells else 0

        if coverage > 0.8 and main_table.header_row is not None and len(merged_cells) < 3:
            return SheetType.TABULAR

        if coverage > 0.4 and main_table.header_row is not None:
            return SheetType.SEMI_STRUCTURED

        return SheetType.UNSTRUCTURED

    def _identify_metadata_regions(
        self,
        cell_matrix: List[List[int]],
        table_regions: List[TableBoundary],
        max_row: int,
        max_col: int,
    ) -> List[Tuple[int, int, int, int]]:
        metadata_regions: List[Tuple[int, int, int, int]] = []

        if not table_regions:
            return metadata_regions

        main_table = table_regions[0]

        if main_table.start_row > 0:
            if self._has_non_empty(cell_matrix, 0, main_table.start_row - 1, 0, max_col - 1):
                metadata_regions.append((0, 0, main_table.start_row - 1, max_col - 1))

        if main_table.end_row < max_row - 1:
            if self._has_non_empty(
                cell_matrix,
                main_table.end_row + 1,
                max_row - 1,
                0,
                max_col - 1,
            ):
                metadata_regions.append((main_table.end_row + 1, 0, max_row - 1, max_col - 1))

        return metadata_regions

    def _has_non_empty(
        self,
        matrix: List[List[int]],
        start_row: int,
        end_row: int,
        start_col: int,
        end_col: int,
    ) -> bool:
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                if matrix[row][col] != 0:
                    return True
        return False
