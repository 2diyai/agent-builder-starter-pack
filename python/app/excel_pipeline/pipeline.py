from io import BytesIO
from typing import Any, Dict, Optional

from openpyxl import load_workbook

from .classifier import SheetClassifier
from .parsers import ParserFactory
from .serializer import JsonFormat, LLMJsonSerializer


class ExcelToLLMPipeline:
    def __init__(
        self,
        json_format: str = JsonFormat.HYBRID,
        max_rows_per_table: Optional[int] = None,
    ):
        self.classifier = SheetClassifier()
        self.serializer = LLMJsonSerializer(
            json_format=json_format,
            max_rows=max_rows_per_table,
        )

    def process_workbook_bytes(self, file_bytes: bytes, source_name: str = "uploaded.xlsx") -> Dict[str, Any]:
        workbook_values = load_workbook(filename=BytesIO(file_bytes), data_only=True)
        workbook_formulas = load_workbook(filename=BytesIO(file_bytes), data_only=False)

        result: Dict[str, Any] = {
            "workbook": source_name,
            "sheet_count": len(workbook_values.sheetnames),
            "sheets": [],
        }

        for sheet_name in workbook_values.sheetnames:
            ws_values = workbook_values[sheet_name]
            ws_formulas = workbook_formulas[sheet_name]

            analysis = self.classifier.analyze(ws_values)
            parser = ParserFactory.get_parser(analysis.sheet_type)
            parsed = parser.parse(ws_values, analysis, ws_formulas)

            serialized = self.serializer.serialize_sheet(parsed)
            serialized["worksheet_name"] = sheet_name
            serialized["sheet_type"] = analysis.sheet_type.value
            serialized["analysis"] = {
                "data_density": analysis.data_density,
                "max_row": analysis.max_row,
                "max_col": analysis.max_col,
                "table_count": len(analysis.table_regions),
                "metadata_region_count": len(analysis.metadata_regions),
                "merged_cells": analysis.merged_cells,
            }
            result["sheets"].append(serialized)

        return result
